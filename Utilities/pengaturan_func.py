from PyQt6 import QtCore, QtWidgets
from PyQt6.QtWidgets import QColorDialog, QMessageBox
from openpyxl.styles import PatternFill
from configparser import ConfigParser
import re
import os

def baca_pengaturan_conf(self):
    config = ConfigParser()
    try:
        # Membaca file konfigurasi
        config.read('pengaturan_app.conf')
        
        # Cek apakah file konfigurasi ada dan bagian yang dibutuhkan ada di dalamnya
        if not config.sections():
            raise FileNotFoundError("File konfigurasi tidak ditemukan, menggunakan nilai default.")

        # Memuat pengaturan suhu
        if 'Pengaturan Suhu' in config:
            batas_suhu_1 = config.getfloat('Pengaturan Suhu', 'Suhu1', fallback=40.0)
            batas_suhu_2 = config.getfloat('Pengaturan Suhu', 'Suhu2', fallback=71.0)
            self.ui.batasSuhu1_DoubleSpinBox.setValue(batas_suhu_1)
            self.ui.batasRangeSuhu2_DoubleSpinBox.setValue(batas_suhu_2)
        else:
            raise KeyError("Bagian 'Pengaturan Suhu' tidak ditemukan di file konfigurasi.")

        # Memuat pengaturan fill warna
        if 'Pengaturan Fill Warna' in config:
            fill_suhu1 = config.get('Pengaturan Fill Warna', 'FillSuhu1', fallback='#00ff00')
            fill_suhu2 = config.get('Pengaturan Fill Warna', 'FillSuhu2', fallback='#ffff00')
            fill_holding_time = config.get('Pengaturan Fill Warna', 'FillHoldingTime', fallback='#ff0000')

            self.ui.pilihWarnaSuhu1_btn.setStyleSheet(f"background-color: {fill_suhu1};")
            self.fill_suhu1 = PatternFill(start_color=fill_suhu1.lstrip('#'),
                                          end_color=fill_suhu1.lstrip('#'),
                                          fill_type="solid")
            self.ui.pilihWarnaSuhu2_btn.setStyleSheet(f"background-color: {fill_suhu2};")
            self.fill_suhu2 = PatternFill(start_color=fill_suhu2.lstrip('#'),
                                          end_color=fill_suhu2.lstrip('#'),
                                          fill_type="solid")
            self.ui.pilihWarnaHoldingTime_btn.setStyleSheet(f"background-color: {fill_holding_time};")
            self.fill_holding_time = PatternFill(start_color=fill_holding_time.lstrip('#'),
                                                 end_color=fill_holding_time.lstrip('#'),
                                                 fill_type="solid")
        else:
            raise KeyError("Bagian 'Pengaturan Fill Warna' tidak ditemukan di file konfigurasi.")

        # Memuat pengaturan holding time
        if 'Pengaturan Holding Time' in config:
            holding_time = config.get('Pengaturan Holding Time', 'HoldingTime', fallback='00:09:00')
            self.ui.holdingTime_lineEdit.setText(holding_time)
        else:
            raise KeyError("Bagian 'Pengaturan Holding Time' tidak ditemukan di file konfigurasi.")

    except FileNotFoundError:
        # Inisialisasi nilai default jika file tidak ditemukan
        self.ui.batasSuhu1_DoubleSpinBox.setValue(40.0)
        self.ui.batasRangeSuhu2_DoubleSpinBox.setValue(71.0)
        self.ui.pilihWarnaSuhu1_btn.setStyleSheet("background-color: #00ff00;")
        self.fill_suhu1 = PatternFill(start_color='00ff00', end_color='00ff00', fill_type="solid")
        self.ui.pilihWarnaSuhu2_btn.setStyleSheet("background-color: #ffff00;")
        self.fill_suhu2 = PatternFill(start_color='ffff00', end_color='ffff00', fill_type="solid")
        self.ui.pilihWarnaHoldingTime_btn.setStyleSheet("background-color: #ff0000;")
        self.fill_holding_time = PatternFill(start_color='ff0000', end_color='ff0000', fill_type="solid")
        self.ui.holdingTime_lineEdit.setText('00:00:00')

    except KeyError as ke:
        QtWidgets.QMessageBox.warning(self, "Error", f"Terjadi kesalahan saat membaca file konfigurasi: {str(ke)}")
    except Exception as e:
        QtWidgets.QMessageBox.critical(self, "Error", f"Terjadi kesalahan: {e}")

def simpan_pengaturan_conf(self):
    try:
        # Ambil nilai dari QDoubleSpinBox
        batas_suhu_1 = self.ui.batasSuhu1_DoubleSpinBox.value()
        batas_suhu_2 = self.ui.batasRangeSuhu2_DoubleSpinBox.value()
        
        # Validasi nilai suhu
        if batas_suhu_1 <= 0 or batas_suhu_2 <= 0:
            raise ValueError("Nilai suhu harus lebih dari 0.")
        
        # Ambil warna suhu dari QPushButton
        warna_suhu1 = self.ui.pilihWarnaSuhu1_btn.styleSheet().split("background-color:")[1].strip().replace(";", "")
        warna_suhu2 = self.ui.pilihWarnaSuhu2_btn.styleSheet().split("background-color:")[1].strip().replace(";", "")
        warna_holding_time = self.ui.pilihWarnaHoldingTime_btn.styleSheet().split("background-color:")[1].strip().replace(";", "")
        
        # Ambil nilai holding time dari QLineEdit
        holding_time = self.ui.holdingTime_lineEdit.text()
        if not holding_time:
            raise ValueError("Holding time tidak boleh kosong.")
        
        # Validasi format holding time (misal: HH:MM:SS)
        if not re.match(r'^\d{1,2}:\d{2}:\d{2}$', holding_time):
            raise ValueError("Format holding time tidak valid. Gunakan format HH:MM:SS.")

        # Simpan nilai ke file konfigurasi
        config = ConfigParser()
        config['Pengaturan Suhu'] = {
            'Suhu1': batas_suhu_1,
            'Suhu2': batas_suhu_2
        }
        config['Pengaturan Fill Warna'] = {
            'FillSuhu1': warna_suhu1,
            'FillSuhu2': warna_suhu2,
            'FillHoldingTime': warna_holding_time
        }
        config['Pengaturan Holding Time'] = {
            'HoldingTime': holding_time
        }
        with open('pengaturan_app.conf', 'w') as configfile:
            config.write(configfile)
        
        # Tampilkan pesan konfirmasi
        QtWidgets.QMessageBox.information(self, "Berhasil Disimpan", "Konfigurasi Pengaturan telah disimpan.")
    except ValueError as ve:
        QtWidgets.QMessageBox.warning(self, "Input Tidak Valid", str(ve))
    except Exception as e:
        QtWidgets.QMessageBox.critical(self, "Error", f"Terjadi kesalahan: {e}")


def get_batas_suhu_1(self):
    return self.ui.batasSuhu1_DoubleSpinBox.value()

def get_batas_suhu_2(self):
    return self.ui.batasRangeSuhu2_DoubleSpinBox.value()

def pilih_warna_suhu1(self):
    color = QColorDialog.getColor()
    if color.isValid():
        hex_color = color.name()  # Mendapatkan warna dalam format hex
        # Set warna yang dipilih ke label
        self.ui.pilihWarnaSuhu1_btn.setStyleSheet(f"background-color: {hex_color};")

def pilih_warna_suhu2(self):
    color = QColorDialog.getColor()
    if color.isValid():
        hex_color = color.name()  # Mendapatkan warna dalam format hex
        # Set warna yang dipilih ke label
        self.ui.pilihWarnaSuhu2_btn.setStyleSheet(f"background-color: {hex_color};")

def pilih_warna_holding_time(self):
    color = QColorDialog.getColor()
    if color.isValid():
        hex_color = color.name()  # Mendapatkan warna dalam format hex
        # Set warna yang dipilih ke label
        self.ui.pilihWarnaHoldingTime_btn.setStyleSheet(f"background-color: {hex_color};")

def baca_pengaturan_folder():
    config = ConfigParser()
    try:
        config.read('pengaturan_app.conf')
        if 'Pengaturan Folder' in config:
            selected_folder = config.get('Pengaturan Folder', 'SelectedFolder', fallback='')
            return selected_folder
        else:
            return ''
    except Exception as e:
        QtWidgets.QMessageBox.critical(None, "Error", f"Terjadi kesalahan saat membaca pengaturan folder: {e}")
        return ''

def simpan_pengaturan_folder(folder):
    try:
        config = ConfigParser()
        config.read('pengaturan_app.conf')
        if not config.has_section('Pengaturan Folder'):
            config.add_section('Pengaturan Folder')
        config.set('Pengaturan Folder', 'SelectedFolder', folder)
        with open('pengaturan_app.conf', 'w') as configfile:
            config.write(configfile)
    except Exception as e:
        QtWidgets.QMessageBox.critical(None, "Error", f"Terjadi kesalahan saat menyimpan pengaturan folder: {e}")
