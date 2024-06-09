import os
import shutil
import openpyxl
from PyQt6 import QtCore, QtWidgets
from datetime import datetime
from openpyxl.styles import PatternFill
from openpyxl.utils.cell import get_column_letter, column_index_from_string
from App.salin_data_waktu_suhu import *
from App.validasi_data_waktu import *
from Utilities.pengaturan_func import *
import win32com.client as win32

def konversi_xls_ke_xlsx(source_file, dest_file):
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    try:
        source_file = os.path.abspath(source_file)
        dest_file = os.path.abspath(dest_file)

        wb = excel.Workbooks.Open(source_file)
        wb.SaveAs(dest_file, FileFormat=51)  # 51 adalah format untuk xlsx
        wb.Close()
        return True, None
    except Exception as e:
        return False, f"Gagal mengonversi {source_file}. Error: {str(e)}"
    finally:
        excel.Quit()

def start_import_excel(self):
    log_folder = "LOG"
    if not os.path.exists(log_folder):
        os.makedirs(log_folder)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file_path = f'{log_folder}/proses_excel_data_log_{timestamp}.txt'

    self.ui.daftarOutputFiles_treeWidget.clear()
    processed_files = []
    failed_files = []

    temp_folder = "konversi_data_temp"
    if not os.path.exists(temp_folder):
        os.makedirs(temp_folder)

    total_files = self.ui.daftarInputFiles_treeWidget.topLevelItemCount()
    self.ui.progressBar.setValue(0)

    for index in range(total_files):
        item = self.ui.daftarInputFiles_treeWidget.topLevelItem(index)
        file_path = item.toolTip(0)

        # Pengecekan format file
        if file_path.endswith('.xls'):
            base_name = os.path.basename(file_path).replace('.xls', '.xlsx')
            dest_file_path = os.path.join(temp_folder, base_name)
            success, error_message = konversi_xls_ke_xlsx(file_path, dest_file_path)
            if success:
                file_path = dest_file_path  # Gunakan file yang sudah dikonversi
            else:
                failed_files.append(error_message)
                QtWidgets.QMessageBox.critical(self, "Error", error_message)
                if not self.confirm_continue():
                    break

        try:
            self.import_excel_with_progress(file_path, failed_files)
            processed_files.append(os.path.basename(file_path))
        except openpyxl.utils.exceptions.InvalidFileException:
            error_message = f"{os.path.basename(file_path)}: Format file tidak didukung, gunakan format Excel .xls atau .xlsx"
            failed_files.append(error_message)
            QtWidgets.QMessageBox.critical(self, "Error", error_message)
            if not self.confirm_continue():
                break
        except Exception as e:
            error_message = f"{os.path.basename(file_path)}: Terjadi kesalahan: {str(e)}"
            failed_files.append(error_message)
            QtWidgets.QMessageBox.critical(self, "Error", error_message)
            if not self.confirm_continue():
                break

        # Mengupdate nilai progress bar setelah setiap file diproses
        progress_value = int(((index + 1) / total_files) * 100)
        self.ui.progressBar.setValue(progress_value)
        QtWidgets.QApplication.processEvents()  # Memproses event agar progress bar terlihat

    with open(log_file_path, 'w') as log_file:
        log_file.write("File Excel yang berhasil diproses:\n")
        for file in processed_files:
            log_file.write(file + '\n')
        log_file.write("\nBaris Excel yang gagal diproses karena DATA TIDAK VALID (bukan angka atau desimal):\n")
        for error in failed_files:
            log_file.write(error + '\n')

    if processed_files:
        QtWidgets.QMessageBox.information(self, 'Berhasil', f'{len(processed_files)} file Excel .xlsx telah disalin dan disesuaikan isi datanya:\n' + '\n'.join(processed_files))

    if failed_files:
        QtWidgets.QMessageBox.warning(self, 'Gagal', f'{len(failed_files)} Total Baris Excel yang gagal diproses karena DATA SUHU bukan angka atau desimal.\nLihat "{log_file_path}" untuk detailnya.')

    # Menghapus folder konversi_data_temp setelah semua proses selesai
    hapus_folder_konversi(temp_folder)

def hapus_folder_konversi(folder_path):
    try:
        shutil.rmtree(folder_path)
    except Exception as e:
        QtWidgets.QMessageBox.critical(None, "Error", f"Gagal menghapus folder {folder_path}. Error: {str(e)}")

def confirm_continue(self):
    reply = QtWidgets.QMessageBox.question(self, "Konfirmasi", 
                                           "Terjadi kesalahan. Apakah Anda ingin melanjutkan proses?",
                                           QtWidgets.QMessageBox.StandardButton.Yes | QtWidgets.QMessageBox.StandardButton.No)
    return reply == QtWidgets.QMessageBox.StandardButton.Yes

def import_excel_with_progress(self, file_path, failed_files):
    temp_folder = "hasil data"
    if not os.path.exists(temp_folder):
        os.makedirs(temp_folder)

    base_name = os.path.basename(file_path)
    temp_file_path = os.path.join(temp_folder, base_name)
    increment = 0
    while os.path.exists(temp_file_path):
        increment += 1
        name, ext = os.path.splitext(base_name)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        temp_file_path = os.path.join(temp_folder, f"{name}_{timestamp}{ext}")
    
    if increment > 0 and not getattr(self, 'auto_overwrite', False):
        reply = QtWidgets.QMessageBox.question(self, "Konfirmasi", 
                                               f"File {base_name} sudah ada. Gunakan nama {os.path.basename(temp_file_path)}?",
                                               QtWidgets.QMessageBox.StandardButton.Yes | QtWidgets.QMessageBox.StandardButton.No | QtWidgets.QMessageBox.StandardButton.YesToAll)
        if reply == QtWidgets.QMessageBox.StandardButton.No:
            return
        elif reply == QtWidgets.QMessageBox.StandardButton.YesToAll:
            self.auto_overwrite = True

    try:
        shutil.copy(file_path, temp_file_path)
    except (PermissionError, OSError) as e:
        QtWidgets.QMessageBox.critical(self, "Error", f"Tidak dapat menyalin file excel karena kamu sedang membuka atau menggunakan filenya atau file tidak ada !! : {str(e)}")
        return

    try:
        wb = openpyxl.load_workbook(temp_file_path)
        ws = wb["Sheet1"]
        ws_data = wb["DATA"]
    except (PermissionError, OSError) as e:
        QtWidgets.QMessageBox.critical(self, "Error", f"Tidak dapat menggunakan file excel karena kamu sedang membuka atau menggunakan filenya atau file tidak ada !!! : {str(e)}")
        return

    # Mengambil informasi file
    file_info = QtCore.QFileInfo(temp_file_path)
    file_name = file_info.fileName()
    file_modified_date = file_info.lastModified().toString(QtCore.Qt.DateFormat.ISODate)
    file_type = file_info.suffix()
    file_size = f"{file_info.size() / 1024:.2f} KB"

    # Membuat item baru dengan informasi file
    item = QtWidgets.QTreeWidgetItem([file_name, file_modified_date, file_type, file_size])
    item.setToolTip(0, temp_file_path)  # Mengatur tooltip untuk item
    self.ui.daftarOutputFiles_treeWidget.addTopLevelItem(item)  # Menambahkan item ke daftarOutputFiles_treeWidget

    # Mengatur ukuran kolom sesuai dengan isi konten setelah semua item ditambahkan
    for i in range(self.ui.daftarOutputFiles_treeWidget.columnCount()):
        self.ui.daftarOutputFiles_treeWidget.resizeColumnToContents(i)

    # Mengatur nilai awal progress bar
    self.ui.progressBar.setValue(0)

    # Menghitung jumlah baris untuk digunakan dalam progress bar
    total_rows = ws.max_row - 3  # Mengurangi header

    # Inisialisasi variabel untuk melacak apakah sedang dalam periode kenaikan suhu di atas 71.0 derajat
    kenaikan_suhu_71 = False

    # Inisialisasi variabel untuk menyimpan nilai suhu dan waktu
    data_suhu_waktu = {
        'kolom_d': {'suhu': [], 'waktu': []},
        'kolom_f': {'suhu': [], 'waktu': []}
    }

    # Mendapatkan nilai dari DoubleSpinBox
    batas_suhu_1 = self.get_batas_suhu_1()
    batas_suhu_2 = self.get_batas_suhu_2()

    # Loop untuk menyalin isi konten kolom H, I, K, dan M
    for i, row in enumerate(range(4, ws.max_row + 1), 1):
        ws[f'B{i + 3}'] = ws[f'H{row}'].value
        ws[f'C{i + 3}'] = ws[f'I{row}'].value
        ws[f'E{i + 3}'] = ws[f'I{row}'].value
        ws[f'D{i + 3}'] = ws[f'K{row}'].value
        ws[f'F{i + 3}'] = ws[f'M{row}'].value

        # Memeriksa suhu di kolom D dan F
        for kolom, kolom_waktu, label in [('D', 'C', 'kolom_d'), ('F', 'E', 'kolom_f')]:
            sel_suhu = ws[f'{kolom}{i + 3}']
            if sel_suhu.value is not None:
                try:
                    suhu = float(sel_suhu.value)
                except ValueError:
                    # Jika nilai tidak dapat dikonversi menjadi float, anggap nilainya None atau tidak ada, log pesan error dan lewati baris ini
                    error_message = f"{os.path.basename(file_path)} - Baris {row}: Nilai data tidak valid untuk konversi ke desimal di kolom {kolom}"
                    failed_files.append(error_message)
                    continue  # Lewati iterasi saat ini jika terjadi kesalahan konversi
                
                # Memeriksa apakah suhu di atas nilai default 71.0 Derajat
                if suhu >= batas_suhu_2:
                    ws[f'{kolom_waktu}{i + 3}'].fill = self.fill_suhu2
                    sel_suhu.fill = self.fill_suhu2
                    kenaikan_suhu_71 = True
                elif suhu < batas_suhu_2 and kenaikan_suhu_71:
                    kenaikan_suhu_71 = False

                # Memeriksa apakah suhu di atas nilai default 40.0 Derajat
                try:
                    if ws[f'{kolom}{i + 2}'].value is not None and isinstance(ws[f'{kolom}{i + 2}'].value, (int, float)):
                        suhu_sebelumnya = float(ws[f'{kolom}{i + 2}'].value)
                    else:
                        suhu_sebelumnya = None
                except ValueError:
                    suhu_sebelumnya = None
                    # Log pesan error jika nilai sebelumnya tidak valid
                    error_message = f"{os.path.basename(file_path)} - Baris {row}: Nilai data sebelumnya tidak valid untuk konversi ke desimal di kolom {kolom}"
                    failed_files.append(error_message)
                    continue  # Lewati iterasi saat ini jika terjadi kesalahan konversi

                # Menandai suhu yang melebihi nilai default 40.0 Derajat sesuai dengan kriteria
                if suhu > batas_suhu_1 and (suhu_sebelumnya is not None and suhu_sebelumnya <= batas_suhu_1):
                    data_suhu_waktu[label]['waktu'].append(ws[f'{kolom_waktu}{i + 3}'].value)
                    data_suhu_waktu[label]['suhu'].append(suhu)
                    ws[f'{kolom_waktu}{i + 3}'].fill = self.fill_suhu1
                    sel_suhu.fill = self.fill_suhu1

        # Mengupdate nilai progress bar di dalam loop
        progress_value = int((i / total_rows) * 100)
        self.ui.progressBar.setValue(progress_value)
        QtWidgets.QApplication.processEvents()  # Memproses event agar progress bar terlihat

    # Menyimpan data suhu dan waktu ke dalam sheet DATA
    for label, indeks_waktu_awal, indeks_suhu_awal in [('kolom_d', 2, 3), ('kolom_f', 4, 5)]:
        indeks_waktu = indeks_waktu_awal
        indeks_suhu = indeks_suhu_awal
        for waktu, suhu in zip(data_suhu_waktu[label]['waktu'], data_suhu_waktu[label]['suhu']):
            ws_data.cell(row=7, column=indeks_waktu).value = waktu
            ws_data.cell(row=7, column=indeks_waktu).fill = self.fill_suhu1
            indeks_waktu += 4
            
            ws_data.cell(row=7, column=indeks_suhu).value = suhu
            ws_data.cell(row=7, column=indeks_suhu).fill = self.fill_suhu1
            indeks_suhu += 4

    # Menjalankan fungsi cek_waktu dengan pembaruan progress bar
    # print("Memanggil cek_waktu...")  # Debug print sebelum pemanggilan cek_waktu
    self.salin_data_waktu_suhu_71_kesheet_DATA(ws, ws_data)
    self.cek_waktu(ws_data, total_rows)
    # print("Selesai memanggil cek_waktu.")  # Debug print setelah pemanggilan cek_waktu

    # Mengupdate nilai progress bar terakhir kali untuk memastikan progress bar penuh
    self.ui.progressBar.setValue(100)
    QtWidgets.QApplication.processEvents()  # Memproses event agar progress bar terlihat

    # Menyimpan perubahan dan menutup file Excel
    wb.save(f"hasil data/{os.path.basename(temp_file_path)}")
    wb.close()


# ===================================================================================