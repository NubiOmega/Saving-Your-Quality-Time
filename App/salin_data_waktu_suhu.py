from openpyxl.utils.cell import get_column_letter, column_index_from_string

# optimalisasi salin data suhu 71 ke datasheet DATA
 
def salin_data_waktu_suhu_71_kesheet_DATA(self, ws, ws_data):
    # Inisialisasi list untuk menyimpan nilai data dari kolom D dan kolom F
    nilai_data_d = []
    nilai_data_f = []

    # Inisialisasi variabel untuk kolom target pertama untuk kolom D dan C
    kolom_target_d = 'C'
    kolom_target_index_d = column_index_from_string(kolom_target_d)
    kolom_target_c = 'B'
    kolom_target_index_c = column_index_from_string(kolom_target_c)

    # Inisialisasi variabel untuk kolom target pertama untuk kolom F dan E
    kolom_target_f = 'E'
    kolom_target_index_f = column_index_from_string(kolom_target_f)
    kolom_target_e = 'D'
    kolom_target_index_e = column_index_from_string(kolom_target_e)

    # Inisialisasi variabel baris untuk menulis data pada sheet target
    baris_d = 15
    baris_f = 15

    # Loop melalui sel di kolom D dan F
    for row in range(1, ws.max_row + 1):
        # Ambil nilai dari sel di kolom D dan F pada baris tertentu
        nilai_d = ws[f'D{row}'].value
        waktu_d = ws[f'C{row}'].value
        nilai_f = ws[f'F{row}'].value
        waktu_f = ws[f'E{row}'].value
        
        # Periksa apakah nilai di kolom D tidak kosong
        if nilai_d is not None:
            # Konversi nilai menjadi tipe data numerik (float) jika memungkinkan
            try:
                nilai_d = float(nilai_d)
            except ValueError:
                continue  # Lewati jika tidak bisa dikonversi
            
            # Periksa apakah nilai >= 71
            if nilai_d >= 71:
                # Tambahkan nilai ke dalam list
                nilai_data_d.append((waktu_d, nilai_d))
            else:
                # Jika nilai <= 71, hentikan proses jika sebelumnya telah menemukan nilai >= 71
                if nilai_data_d:
                    # Salin nilai_data ke cell pada kolom target untuk kolom D
                    for waktu, nilai in nilai_data_d:
                        # Salin nilai waktu dari kolom C ke cell pada kolom target untuk kolom C
                        cell_c = ws_data[f'{kolom_target_c}{baris_d}']
                        cell_c.value = waktu
                        
                        cell_d = ws_data[f'{kolom_target_d}{baris_d}']
                        cell_d.value = nilai
                        
                        # Pindah ke baris berikutnya
                        baris_d += 1
                    
                    # Kosongkan nilai_data untuk kolom D untuk menampung nilai data berikutnya
                    nilai_data_d = []
                    
                    # Pindah ke kolom target berikutnya untuk kolom D dan C (menggunakan langkah 4 kolom)
                    kolom_target_index_d += 4
                    kolom_target_d = get_column_letter(kolom_target_index_d)
                    kolom_target_index_c += 4
                    kolom_target_c = get_column_letter(kolom_target_index_c)
                    
                    # Reset baris_d ke nilai awal
                    baris_d = 15
        
        # Periksa apakah nilai di kolom F tidak kosong
        if nilai_f is not None:
            # Konversi nilai menjadi tipe data numerik (float) jika memungkinkan
            try:
                nilai_f = float(nilai_f)
            except ValueError:
                continue  # Lewati jika tidak bisa dikonversi
            
            # Periksa apakah nilai >= 71
            if nilai_f >= 71:
                # Tambahkan nilai ke dalam list
                nilai_data_f.append((waktu_f, nilai_f))
            else:
                # Jika nilai <= 71, hentikan proses jika sebelumnya telah menemukan nilai >= 71
                if nilai_data_f:
                    # Salin nilai_data ke cell pada kolom target untuk kolom F
                    for waktu, nilai in nilai_data_f:
                        # Salin nilai waktu dari kolom E ke cell pada kolom target untuk kolom E
                        cell_e = ws_data[f'{kolom_target_e}{baris_f}']
                        cell_e.value = waktu
                        
                        cell_f = ws_data[f'{kolom_target_f}{baris_f}']
                        cell_f.value = nilai
                        
                        # Pindah ke baris berikutnya
                        baris_f += 1
                    
                    # Kosongkan nilai_data untuk kolom F untuk menampung nilai data berikutnya
                    nilai_data_f = []
                    
                    # Pindah ke kolom target berikutnya untuk kolom F dan E (menggunakan langkah 4 kolom)
                    kolom_target_index_f += 4
                    kolom_target_f = get_column_letter(kolom_target_index_f)
                    kolom_target_index_e += 4
                    kolom_target_e = get_column_letter(kolom_target_index_e)
                    
                    # Reset baris_f ke nilai awal
                    baris_f = 15
        
        # Jika sudah mencapai kolom terakhir, keluar dari loop
        if kolom_target_index_d > column_index_from_string('BQ') and kolom_target_index_f > column_index_from_string('BQ'):
            break

# ===================================================================================